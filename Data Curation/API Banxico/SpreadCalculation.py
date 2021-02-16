
# Fixed Income Spreads

'The objective of this code is the analysis for bonos spreads using Banxico API'

# Libraries

from sympy import *
from scipy.optimize import *
import numpy as np
from pathlib import Path
from datetime import date
import pathlib
from datetime import datetime, timedelta
import pandas as pd
import requests
import openpyxl
import mpmath
from scipy.optimize import broyden1
rutac=pathlib.Path(__file__).parent.absolute()
rutac=str(rutac)
excel_path = Path( rutac+'\FI Series List.xlsx')
wb = openpyxl.load_workbook(excel_path)
ws = wb["Bonos"]
today_date=datetime.today()
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker
import matplotlib

# Scaling Datasets

def spreadb(arr):
    global pd1n,pd2n,pd1,pd2,result
    result= pd.DataFrame(data={})
    for i in range(len(arr)):
        pd1=pd.read_excel(open('FI Series List.xlsx', 'rb'),
              sheet_name=str(arr[i]))

        pd1=pd1.set_index(pd1["Fecha"])

        for j in range(len(arr)):
            if arr[i]!=arr[j]:
                pd2=pd.read_excel(open('FI Series List.xlsx', 'rb'),
                              sheet_name=str(arr[j]))
                
                pd2=pd2.set_index(pd2["Fecha"])

                result["Data1"]=pd1["r"]-pd2["r"]
                result["Media"]=result["Data1"].mean()
                result["Std+"]=result["Media"]+2*result["Data1"].std()
                result["Std-"]=result["Media"]-2*result["Data1"].std()
                fig, ax1 = plt.subplots()
                plt.title(arr[i]+ " y "+ arr[j])
                color = '#006d91'                
                ax1.set_xlabel('Observations')
                ax1.set_ylabel('Spread', color=color)
                color = '#006d91'
                ax1.plot(result["Data1"], color=color)
                color = '#575759'
                ax1.plot(result["Media"], color=color)
                
                color = '#8c5150'
                ax1.plot(result["Std+"], color=color)
                color = '#8c5150'
                ax1.plot(result["Std-"], color=color)               
                ax1.tick_params(axis='y')
                
                ax1.xaxis.set_major_locator(ticker.LinearLocator(4))
                
                fig.tight_layout()  # otherwise the right y-label is slightly clipped
                plt.show()

token="09decb94a9f6a633242c5e59b440e7ad1174143665e653c8d48ad4e87be1d4cb" 

#The token and fuction to download data from Banxico's API         

def descargardatosBX(serie,fechaini,fechafin,token):
    #Receiving and using the series, dates and token key
    url="https://www.banxico.org.mx/SieAPIRest/service/v1/series/"+serie+"/datos/"+fechaini+"/"+fechafin
    headers={'Bmx-Token':token}
    response=requests.get(url,headers=headers)
    status=response.status_code
    if status!=200:
        
        return print("Error,codigo {}".format(status))
    raw_data=response.json()
    data=raw_data['bmx']['series'][0]['datos']
    df=pd.DataFrame(data)
    #saving and cleaning the daata from the series
    df["dato"]=df["dato"].str.replace(',','')
    df["dato"]=df["dato"].str.replace('N/E','0')
    df['dato']=df['dato'].apply(lambda x:float(x))
    df['fecha']=pd.to_datetime(df['fecha'],format='%d/%m/%Y')
    #df.set_index('fecha',inplace=True)
    return df


# Reads an excel

dfb=pd.read_excel("FI Series List.xlsx")

for i in range(len(dfb["Serieplimp"])):

    #with the data and series we loop in the bonds
    if dfb["Bono"][i]!="Bono 2-30":
        spzo=dfb["SeriePlazo"][i]
        splimpio=dfb["Serieplimp"][i]
        scup=dfb["Seriecup"][i]
        
        fi="2020-01-01"
        fin="2021-02-16"
        #Range of work
        
        # fi="2021-09-02"
        # fin="2021-12-02"
            
        print(spzo)
        #Downloading data from banxico
        datapzo=descargardatosBX(spzo,fi,fin,token)
        datapl=descargardatosBX(splimpio,fi,fin,token)
        datacup=descargardatosBX(scup,fi,fin,token)      
        datapzo.to_excel("plaz.xlsx")  
        
        
        datasize=(len(datapl["dato"])-1)
        import QuantLib as ql
        import pandas as pd
        #Using quantlib to manipulate dates easily and splitting string dates

        #Looping in the data individually per bond|
        for j in range(len(datapl["dato"])-datasize+10):
            if j==0 :
                ws = wb["Bonos"]  
                ws.cell(row=2+j+i, column=5).value=datapl["fecha"][datasize-j]+ timedelta(datapzo["dato"][datasize-j])
                diai=int(dfb["Day"][i])
                messi=int(dfb["Month"][i])
                ani=int(dfb["Year"][i])
                matur=str(datapl["fecha"][datasize-j]+ timedelta(datapzo["dato"][datasize-j])).split()
                mbon=matur[0].split("-")
                diaf=int(mbon[2])
                mesf=int(mbon[1])
                anf=int(mbon[0])
            date = ql.Date(diai,messi , ani)-j
            raw_date=ql.Date(diaf, mesf, anf)-j
            print(date, " - ",raw_date)            
            inicio=date-raw_date
            mx_calendar = ql.Mexico()
            #Using a date we know of settlement we use it to obtain the bussiness daysdecreasing by one the date
            mx_busdays = mx_calendar.businessDaysBetween(date, raw_date)
           
            VN=100
            TC=(datacup["dato"][datasize-j])/100
            d=mx_busdays#
            dk=datapzo["dato"][datasize-j]-5
        
            K=round(d/182,0)
            
            diastr=182-(dk%182)
            p=round(datapl["dato"][datasize-j],5)
            c=VN*182*TC/360
            r = Symbol('r')
                 
            #Declaring R and clearing from the formula of fixed Bonds we can get our R
            soln=nsolve(((c)+(c)*((1/r)-(1/(r*((1+r)**(K-1)))))+(VN/((1+r)**(K-1))))/((1+r)**(1-((diastr)/182)))+(-(c*diastr/182))-p,.01)
            print("Bono","  ","DiasVencer ","Tasa Cupon","Precio limpio","  Dias trans")
            print(dfb["Bono"][i]," ",dk,"   ",TC,"      ",p, diastr)
            print("Tasa r: ",soln*36000/182)
        
        #depending on how many data you have you can change the days
            rmin=soln*36000/182
            ws = wb[dfb["Bono"][i]]  
            ws.cell(row=2+j, column=1).value=datapl["fecha"][datasize-j]
            ws.cell(row=2+j, column=2).value=dfb["Bono"][i]
            ws.cell(row=2+j, column=3).value=d
            ws.cell(row=2+j, column=4).value=dk
            ws.cell(row=2+j, column=5).value=p
            ws.cell(row=j+2, column=6).value=float(rmin)
            ws.cell(row=j+2, column=7).value=str(matur)
            wb.save(excel_path)
            
arr=["Bono 0-3","Bono 3-5","Bono 7-10","Bono 10-20"]
spreadb(arr)

# mx_busdays =879
# VN=100
# TC=.0675
# d=mx_busdays
# dk=751
# K=round(d/182,0)
# diastr=182-(dk%182)
# p=105.2505
# c=VN*182*TC/360
# r = Symbol('r')
     
# soln=nsolve(((c)+(c)*((1/r)-(1/(r*((1+r)**(K-1)))))+(VN/((1+r)**(K-1))))/((1+r)**(1-((diastr)/182)))+(-(c*diastr/182))-p,.001)
# print(soln*36000/182)