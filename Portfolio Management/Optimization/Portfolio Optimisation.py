
# Libraries

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import mlfinlab as ml
from mlfinlab.portfolio_optimization import MeanVarianceOptimisation
import pandas_datareader as dr
from math import sqrt
from sklearn.cluster import KMeans
from matplotlib import pyplot as plt
from numpy.core.shape_base import vstack
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.cluster.vq import kmeans,vq
import seaborn as sns
import openpyxl
from pathlib import Path
import pathlib

# Read raw data and current portfolio inforamtion

rutac=pathlib.Path(__file__).parent.absolute()
rutac=str(rutac)
excel_path = Path( rutac+'\porfolio.xlsx')
wb = openpyxl.load_workbook(excel_path)
ws = wb["porfolio"]
ws.cell(row=1, column=1).value=ws.cell(row=1, column=1).value
wb.save(excel_path)
def tomaval(columna,ultimop,wb,ws,rutac):
    
    columexc=np.arange(30)
    er=[]

    filaexc=np.arange(len(ultimop)+2)
    print(filaexc)
    for i in range(len(columexc)):
        if i>0:
            if ws.cell(row=1, column=i).value=="Price":
                pcol=i
            if ws.cell(row=1, column=i).value=="Target Price":
                Tpcol=i
                
            if ws.cell(row=1, column=i).value==columna:
                for j in range(len(filaexc)):
                    if j>0:
                        try:
                            tic= ws.cell(row=j, column=2).value
                            Tprice=ws.cell(row=j, column=Tpcol).value
                            pric=ws.cell(row=j, column=pcol).value
                            eret=(Tprice/pric)-1
                            er.append(eret)

                        except:
                            print('Print ticker to detect any issue: ', tic)
                    else:
                        pass
        else:
            print('Print row to detect any issue: ', i)
    df = pd.DataFrame(er,columns=['Expected returns'])
    print(df)
    return df
'Read tickers'
tickers = pd.read_excel('porfolio.xlsx')['Ticker'].dropna()
tickers.drop(tickers.tail(1).index, inplace = True)

expected_returns = pd.read_excel('porfolio.xlsx')['Expected Return'].dropna()
expected_returns = tomaval("Expected Return",tickers,wb,ws,rutac)

l_bounds = pd.read_excel('porfolio.xlsx')['Lower Bound'].dropna()
u_bounds = pd.read_excel('porfolio.xlsx')['Upper Bound'].dropna()

bounds = [ ]
for i in range(len(u_bounds)):
    bounds.append('weights['+str(i)+']>='+str(l_bounds[i]))
    bounds.append('weights['+str(i)+']<='+str(u_bounds[i]))
rutac=pathlib.Path(__file__).parent.absolute()
rutac=str(rutac)
excel_path = Path( rutac+'\stock_prices.xlsx')
wb = openpyxl.load_workbook(excel_path)
ws = wb["Sheet1"]
ws.cell(row=1, column=1).value=ws.cell(row=1, column=1).value
wb.save(excel_path)
wb.close()
'Read tickers'
'Read prices'
stock_prices = pd.read_excel('stock_prices.xlsx')
stock_prices = stock_prices.set_index('Date')

'Generate returns matrix'
ln_matrix = pd.DataFrame(data={})                             
for k in range(len(tickers)):
    ln_matrix[tickers[k]] = np.diff(np.log(stock_prices[tickers[k]]))

# Covariance matrix

print('\nDe-Noising Covariance Matrix\n')

'Simple Covariance'
cov_matrix = ln_matrix.cov()
plt.figure(figsize=(15, 15))
sns.heatmap(cov_matrix, annot=False, cmap='PuOr')
plt.title('Simple Covariance')
plt.show()

'Setting the required parameters for de-noising'

'A class that has the Minimum Covariance Determinant estimator'
risk_estimators = ml.portfolio_optimization.RiskEstimators()

'Relation of number of observations T to the number of variables N (T/N)'
tn_relation = stock_prices.shape[0] / stock_prices.shape[1]

'The bandwidth of the KDE kernel'
kde_bwidth = 0.25

'Finding the De-noised Сovariance matrix'
cov_matrix_denoised = risk_estimators.denoise_covariance(cov_matrix, tn_relation, kde_bwidth)

'Transforming De-noised Covariance from np.array to pd.DataFrame'
cov_matrix_denoised = pd.DataFrame(cov_matrix_denoised, index=cov_matrix.index, columns=cov_matrix.columns)

plt.figure(figsize=(15, 15))
sns.heatmap(cov_matrix_denoised, annot=False, cmap='PuOr')
plt.title('Denoised Covariance')
plt.show()

# inverse_variance 

"""
In statistics, inverse-variance weighting is a method of aggregating two or more random variables to minimize
the variance of the weighted average. Each random variable is weighted in inverse proportion to its variance,
i.e. proportional to its precision.
"""

'Creating our portfolio weights under the correct objective function'
mvoIV = MeanVarianceOptimisation()
mvoIV.allocate(asset_names=stock_prices.columns,
               asset_prices=stock_prices,
               covariance_matrix =cov_matrix,
               expected_asset_returns = expected_returns,
               solution='inverse_variance')

'plotting our optimal portfolio'
IV_weights = mvoIV.weights
y_pos = np.arange(len(IV_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(IV_weights.columns), IV_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Inverse-Variance Portfolio', size=12)
plt.show()

print('\n Inverse-variance results: \n')
port_metrics = mvoIV.get_portfolio_metrics()

# min_volatility 

"""
For this solution, the objective is to generate a portfolio with the least variance. 
"""

'Creating our portfolio weights under the correct objective function'
mvoMV = MeanVarianceOptimisation()
mvoMV.allocate(asset_names=stock_prices.columns,
               asset_prices=stock_prices,
               covariance_matrix =cov_matrix,
               expected_asset_returns = expected_returns,
               solution='min_volatility')

'plotting our optimal portfolio'
MV_weights = mvoMV.weights
y_pos = np.arange(len(MV_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(MV_weights.columns), MV_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Minimum Volatility Portfolio', size=12)
plt.show()

print('\n Minimum Volatility results: \n')
port_metrics = mvoIV.get_portfolio_metrics()

# max_sharpe

"""
For this solution, the objective is (as the name suggests) to maximise the Sharpe Ratio. 
"""

'creating our portfolio weights under the correct objective function'
mvoMS = MeanVarianceOptimisation()
mvoMS.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='max_sharpe')

'plotting our optimal portfolio'
MS_weights = mvoMS.weights
y_pos = np.arange(len(MS_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(MS_weights.columns), MS_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Maximum Sharpe Portfolio', size=12)
plt.show()

print('\n Maximum Sharpe results: \n')
port_metrics = mvoMS.get_portfolio_metrics()

# efficient_risk

"""
For this solution, the objective is to minimise risk given a target return value by the investor. 
"""

target_return = 0.2

'creating our portfolio weights under the correct objective function'
mvoER = MeanVarianceOptimisation()
mvoER.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='efficient_risk',
                target_return=target_return)

'plotting our optimal portfolio'
ER_weights = mvoER.weights
y_pos = np.arange(len(ER_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(ER_weights.columns), ER_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Efficient Risk Portfolio', size=12)
plt.show()

print('\n Efficient Risk results: \n')
port_metrics = mvoER.get_portfolio_metrics()

# efficient_return 

"""
For this solution, the objective is to maximise the portfolio return given a target risk value by the investor. 
"""

target_risk  = 0.0003

'creating our portfolio weights under the correct objective function'
mvoERe = MeanVarianceOptimisation()
mvoERe.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='efficient_return',
                target_risk=target_risk)

'plotting our optimal portfolio'
ERe_weights = mvoERe.weights
y_pos = np.arange(len(ERe_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(ERe_weights.columns), ERe_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Efficient Return Portfolio', size=12)
plt.show()

print('\n Efficient Return results: \n')
port_metrics = mvoERe.get_portfolio_metrics()

# max_return_min_volatility

"""
This is often referred to as quadratic risk utility. The objective function consists of both the 
portfolio return and the risk.
"""

target_return = 0.05
target_risk  = 0.0003

'creating our portfolio weights under the correct objective function'
mvoMRMV = MeanVarianceOptimisation()
mvoMRMV.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='max_return_min_volatility',
                target_return=target_return,
                target_risk=target_risk)

'plotting our optimal portfolio'
MRMV_weights = mvoMRMV.weights
y_pos = np.arange(len(MRMV_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(MRMV_weights.columns), MRMV_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Maximum Return - Minimum Volatility Portfolio', size=12)
plt.show()

print('\n Maximum Return - Minimum Volatility results: \n')
port_metrics = mvoMRMV.get_portfolio_metrics()

# max_diversification

"""
Maximum diversification portfolio tries to diversify the holdings across as many assets as possible.
"""

'creating our portfolio weights under the correct objective function'
mvoMD = MeanVarianceOptimisation()
mvoMD.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='max_diversification')

'plotting our optimal portfolio'
MD_weights = mvoMD.weights
y_pos = np.arange(len(MD_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(MD_weights.columns), MD_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Maximum Diversification Portfolio', size=12)
plt.show()

print('\n Maximum Diversification results: \n')
port_metrics = mvoMD.get_portfolio_metrics()

# max_decorrelation

"""
For this solution, the objective is to minimise the correlation between the assets of a portfolio.
"""

'creating our portfolio weights under the correct objective function'
mvoMDe = MeanVarianceOptimisation()
mvoMDe.allocate(asset_names=stock_prices.columns,
                asset_prices=stock_prices,
                covariance_matrix =cov_matrix,
                expected_asset_returns = expected_returns,
                solution='max_decorrelation')

'plotting our optimal portfolio'
MDe_weights = mvoMDe.weights
y_pos = np.arange(len(MDe_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(MDe_weights.columns), MDe_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Maximum Decorrelation Portfolio', size=12)
plt.show()

print('\n Maximum Decorrelation results: \n')
port_metrics = mvoMDe.get_portfolio_metrics()

# Custom MVO Portfolio - Efficient Risk Solution with Cov Matrix

"""
'Importing ReturnsEstimation class from MLFinLab'
from mlfinlab.portfolio_optimization import ReturnsEstimators

'Calculating our asset returns in order to calculate our covariance matrix'
returns = ReturnsEstimators.calculate_returns(stock_prices)

'Calculating our covariance matrix'
cov = returns.cov()

'Calculating our mean asset returns'
mean_returns = ReturnsEstimators.calculate_mean_historical_returns(stock_prices)
mean_returns
"""

'From here, we can now create our portfolio, adding de-noised cov matrix'
mvo_custom = MeanVarianceOptimisation()
mvo_custom.allocate(asset_names=stock_prices.columns,
                    expected_asset_returns=expected_returns,
                    covariance_matrix=cov_matrix_denoised,
                    target_return = 0.2,
                    solution='efficient_risk')

'plotting our optimal portfolio'
custom_weights = mvo_custom.weights
y_pos = np.arange(len(custom_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(custom_weights.columns), custom_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Custom MVO Portfolio - Efficient Risk Solution with Cov Matrix', size=12)
plt.show()

print('\n Custom MVO Portfolio - Efficient Risk Solution with Cov Matrix results: \n')
port_metrics = mvo_custom.get_portfolio_metrics()

# Custom MVO Portfolio - Efficient Risk Solution with Weight Bounds

'Adding bounds'
mvo_custom_bounds = MeanVarianceOptimisation()
mvo_custom_bounds.allocate(asset_prices=stock_prices,
                           weight_bounds=bounds,
                           solution='max_decorrelation')

'plotting our optimal portfolio'
custom_bounds_weights = mvo_custom_bounds.weights
y_pos = np.arange(len(custom_bounds_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(custom_bounds_weights.columns), custom_bounds_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Custom MVO Portfolio - Efficient Risk Solution with Weight Bounds', size=12)
plt.tight_layout()
'save last chart'
plt.savefig("mvo_custom_weight_bounds.png", dpi=150)
plt.show()

print('\n Custom MVO Portfolio - Efficient Risk Solution with Weight Bounds results: \n')
port_metrics = mvo_custom_bounds.get_portfolio_metrics()

# Custom Portfolio with Custom Objective Function

non_cvxpy_variables = {
    'asset_prices': stock_prices,
    'num_assets': stock_prices.shape[1],
    'covariance': stock_prices.cov(),
    'asset_names': stock_prices.columns,
    'expected_returns': expected_returns
}

cvxpy_variables = [
        'risk = cp.quad_form(weights, covariance)',
        'portfolio_return = cp.matmul(weights, expected_returns)'
]

custom_obj = 'cp.Minimize(risk)'

constraints = ['cp.sum(weights) == 1', 'weights >= 0', 'weights <= 1']

mvo_custom_portfolio = MeanVarianceOptimisation()
mvo_custom_portfolio.allocate_custom_objective(non_cvxpy_variables=non_cvxpy_variables,
                                               cvxpy_variables=cvxpy_variables,
                                               objective_function=custom_obj,
                                               constraints=constraints)

'plotting our optimal portfolio'
mvo_custom_weights = mvo_custom_portfolio.weights
y_pos = np.arange(len(mvo_custom_weights.columns))

plt.rcParams.update(plt.rcParamsDefault)
plt.figure(figsize=(2.5,20))
plt.barh(list(mvo_custom_weights.columns), mvo_custom_weights.values[0])
plt.yticks(y_pos, size=10)
plt.ylabel('Assets', size=12)
plt.xlabel('Asset Weights', size=12)
plt.title('Custom Portfolio with Custom Objective Function', size=12)
plt.show()

print('\n Custom Portfolio with Custom Objective Functio results: \n')
port_metrics = mvo_custom_portfolio.get_portfolio_metrics()

# Plotting

mvo = MeanVarianceOptimisation()
covariance = cov_matrix_denoised
plot = mvo.plot_efficient_frontier(covariance=covariance,
                                   max_return=1.0,
                                   expected_asset_returns=expected_returns)

# Reference
'https://hudsonthames.org/portfolio-optimisation-with-mlfinlab-mean-variance-optimisation/'
'https://mlfinlab.readthedocs.io/en/latest/portfolio_optimisation/mean_variance.html#'